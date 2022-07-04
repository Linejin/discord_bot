#run.py
import discord
import asyncio
from discord.ext import commands

from openpyxl import Workbook
import datetime as dt

import requests

discord.MemberCacheFlags.all()
intents = discord.Intents.default()
intents.members = True
prefix_word = '*'
# client = discord.Client()
bot = commands.Bot(command_prefix=prefix_word, intents = intents)
soundfile_cnt = 0
channelList = []
def is_channel(ctx):   
    return ctx.channel.id in channelList

def is_admin(ctx):   
    for role in ctx.author.roles:
        if "관리자⛑" == role.name:
            return True
    return False

@bot.event
async def on_ready():
    channelList.append(939009875877441587)
    channelList.append(932576925220757566)
    channelList.append(944042644726759424)
    channelList.append(955873043052363798)
   
@bot.event
async def on_command_error(ctx, error):
    if not is_channel(ctx):
        return
    if isinstance(error, commands.CommandNotFound):
    	await ctx.send("CommandNotFound, Invalid Command")

@bot.command(aliases = ["made"])
async def test_made(ctx):
    if not is_channel(ctx):
        return
    await ctx.send("Made by Sejin")

@bot.command()
async def print(ctx):
    if not is_channel(ctx):
        return
    
    member_list = ''
    for member in ctx.message.guild.members:
        member_list += member.name + "\n"
    await ctx.send(member_list) 

@bot.command()
async def leave(ctx): # Note: ?leave won't work, only ?~ will work unless you change  `name = ["~"]` to `aliases = ["~"]` so both can work.
    if (ctx.voice_client): # If the bot is in a voice channel 
        await ctx.guild.voice_client.disconnect() # Leave the channel
    else: # But if it isn't
        await ctx.send("I'm not in a voice channel, use the join command to make me join")

@bot.command()
async def command(ctx):
    await ctx.send(
    f'''-----------------------------------
command without requirements.
{prefix_word}join : join to voice channel.
{prefix_word}leave : leave to voice channel.
{prefix_word}t (message) : (message) To Speak. TTS.  
-----------------------------------
command required admin channel.
{prefix_word}MemberList : all member list in server. required to admin channel.
{prefix_word}noneRoleMember : some of member list who didn\'t check role'''
    )

@bot.command()
async def join(ctx):    
    await ctx.send("init join")
    if ctx.author.voice and ctx.author.voice.channel:
        if ctx.voice_client and ctx.author.voice.channel.id == ctx.voice_client.channel.id:
            await ctx.send("Already joined")
        elif ctx.voice_client and ctx.author.voice.channel.id != ctx.voice_client.channel.id:
            await leave(ctx)
        channel = ctx.author.voice.channel
        await channel.connect()
    else:
        await ctx.send("Can't find channel")

async def makeAllMemberList(ctx, filter = [], pr = ""):
    member_list_index = ['display_name', 'name', 'discriminator', 'joined_at', 'role']
    server_list = ['루페온','카단','실리안','카제로스','아만','카마인','아브렐슈드','니나브']
    member_list = []
    for member in ctx.message.guild.members:
        isrole = False
        for role in member.roles:
            if role.name in filter:
                isrole = True
        if isrole:
            continue
        el = {}
        el['display_name'] = member.display_name
        el['name'] = member.name
        el['discriminator'] = member.discriminator
        el['joined_at'] = member.joined_at.strftime("%Y.%m.%d-%H:%M:%S")
        el['role'] = ""
        for role in member.roles:
            if role.name in server_list:
                el['role'] = role.name
        # el['role'] = []
        # for role in member.roles:
        #     el['role'].append(role.name)
        member_list.append(el)
    if pr == "server" :
        member_list.sort(key = lambda x : (x['role'], x['display_name']))
    elif pr == "join" :
        member_list.sort(key = lambda x : x['joined_at'])
    else:
        member_list.sort(key = lambda x : x['display_name'])
    return member_list

@bot.command()
async def MemberList(ctx, *args):
    if not is_channel(ctx):
        return
    pr = ""
    if not (args is None):
        pr = "".join(args).strip()
    path = await makeMemberList(ctx, filter = ['봇'], pr = pr)
    
    await ctx.send(file=discord.File(path))

@bot.command()
async def noneRoleMember(ctx, *args):
    if not is_channel(ctx):
        return
    pr = "".join(args).strip()
    path = await makeMemberList(ctx, filter = ['남자', '여자', '봇'], pr = pr)
    await ctx.send(file=discord.File(path))

async def makeMemberList(ctx, filter = [] , pr = ""):
    wb = Workbook()
    member_list_index = ['display_name', 'name', 'discriminator', 'joined_at', 'role']
    member_list = await makeAllMemberList(ctx, filter, pr)
    ws = wb.active
    ws.append(member_list_index)
    for member in member_list:
        el = []
        el.append(member['display_name'])
        el.append(member['name'])
        el.append(member['discriminator'])
        el.append(member['joined_at'])
        el.append(member['role'])
        ws.append(el)
    nowDate = dt.datetime.now()
    nowdataStr = nowDate.strftime("%Y %m %d-%H %M %S")
    wb.save("./data/"+ nowdataStr +".xlsx")
    return "./data/"+ nowdataStr +".xlsx"
    
@bot.command(pass_context=True, aliases = ["t"])
async def tts(ctx, *args):
    if not is_channel(ctx):
        return
    if not(ctx.author.voice and ctx.author.voice.channel):
        return
    for member in ctx.author.voice.channel.members:
        if member.id == 958894633654583357:
            return

    if not ctx.voice_client:
        await join(ctx)
    if ctx.author.voice.channel.id != ctx.voice_client.channel.id:
        if len(ctx.voice_client.channel.members) < 2:
            await join(ctx)
        else:
            return
    if ctx.voice_client.is_playing():
        await ctx.send("아직 봇이 말하고 있어요!")
        return

    KAKAO_SECRET_KEY = '8fe999d88a71611a14ad568513311af8'
    headers = {
        #Transfer-Encoding: chunked # 보내는 양을 모를 때 헤더에 포함한다.
        'Host': 'kakaoi-newtone-openapi.kakao.com',
        'Content-Type': 'application/xml',
        'X-DSS-Service': 'DICTATION',
        'Authorization': f'KakaoAK {KAKAO_SECRET_KEY}',
    }
    voiceType = '"WOMAN_READ_CALM"'
    for role in ctx.author.roles:
        if "여자" == role.name:
            voiceType = '"WOMAN_READ_CALM"'
        elif "남자" == role.name:
            voiceType = '"MAN_READ_CALM"'
    data = '<speak><voice name=' + voiceType + '><prosody rate="slow" volume="soft">' + ' '.join(args) +'</prosody></voice></speak>'
    response = requests.post('https://kakaoi-newtone-openapi.kakao.com/v1/synthesize', headers=headers, data=data.encode('utf-8'))
    # 요청 URL과 headers, data를 post방식으로 보내준다.

    rescode = response.status_code
    print(rescode)
    if(rescode==200):
        response_body = response.content
        file_name = ""
        global soundfile_cnt
        while True:
            try :
                file_name = './soundfile/tts' + str(soundfile_cnt) + '.mp3'
                #file_name = './tts' + str(soundfile_cnt) + '.mp3'
                with open(file_name, 'wb') as f:
                    f.write(response_body)
                break
            except Exception as e :
                soundfile_cnt += 1
        await play_soundfile(ctx, file_name)
    else:
        ctx.send("TTS API Error Code:" + rescode)

tts_volume = 1.0
async def play_soundfile(ctx, file_name):
        guild = ctx.guild
        voice_client: discord.VoiceClient = discord.utils.get(bot.voice_clients, guild=guild)
        audio_source = discord.FFmpegPCMAudio(file_name)
        if not voice_client.is_playing():
            voice_client.play(audio_source, after=None)
        else :
            await ctx.send("아직 봇이 말하고 있어요!")

@bot.command()
async def MissionList(ctx):
    if not is_admin(ctx):
        await ctx.message.delete()
        return
    CH = ctx.channel
    for channel in ctx.guild.channels:
        if channel.id == 955681106512973824 or channel.id == 944042644726759424:
            CH = channel
    wb = Workbook()
    member_list_index = ['display_name', 'name', 'discriminator', 'message']
    ws = wb.active
    ws.append(member_list_index)
    await ctx.message.delete()
    li = []
    nowDate = dt.datetime.now()
    threshold = nowDate - dt.timedelta(days=15)
    async for message in CH.history(after = threshold, oldest_first = False):
        member = message.author
        el = {}
        el['display_name'] = member.display_name
        el['name'] = member.name
        el['discriminator'] = member.discriminator
        el['message'] = message.content
        el['created_at'] = message.created_at
        li.append(el)
    li.sort(reverse = True, key = lambda x : (x['display_name'], x['created_at']))
    used = []
    for member in li:
        el = []
        el.append(member['display_name'])
        el.append(member['name'])
        el.append(member['discriminator'])
        el.append(member['message'])
        if member['display_name'] in used:
            continue
        used.append(member['display_name'])
        ws.append(el)
    nowDate = dt.datetime.now()
    nowdataStr = nowDate.strftime("Mission %Y %m %d-%H %M %S")
    wb.save("./data/"+ nowdataStr +".xlsx")
    path = "./data/"+ nowdataStr +".xlsx"
    temp = await ctx.send(file=discord.File(path))
    await temp.delete(delay = 10)

@bot.command()
async def MissionFailList(ctx):
    if not is_admin(ctx):
        await ctx.message.delete()
        return
    CH = ctx.channel
    for channel in ctx.guild.channels:
        if channel.id == 955681106512973824 or channel.id == 944042644726759424:
            CH = channel
    wb = Workbook()
    member_list_index = ['display_name', 'name', 'discriminator', 'joined_at', 'role']
    ws = wb.active
    ws.append(member_list_index)
    await ctx.message.delete()
    li = []
    nowDate = dt.datetime.now()
    threshold = nowDate - dt.timedelta(days=15)
    async for message in CH.history(after = threshold, oldest_first = False):
        member = message.author
        el = {}
        el['display_name'] = member.display_name
        el['name'] = member.name
        el['discriminator'] = member.discriminator
        el['message'] = message.content
        el['created_at'] = message.created_at
        li.append(el)
    li.sort(reverse = True, key = lambda x : (x['display_name'], x['created_at']))

    member_list = await makeAllMemberList(ctx, filter = ["봇", "관리자⛑"], pr = "server")
    for member in member_list:
        flag = False
        for temp in li:
            if temp['display_name'] == member['display_name']:
                flag = True
                break
        if flag :
            continue
        el = []
        el.append(member['display_name'])
        el.append(member['name'])
        el.append(member['discriminator'])
        el.append(member['joined_at'])
        el.append(member['role'])
        ws.append(el)
    nowDate = dt.datetime.now()
    nowdataStr = nowDate.strftime("Fail %Y %m %d-%H %M %S")
    wb.save("./data/"+ nowdataStr +".xlsx")
    path = "./data/"+ nowdataStr +".xlsx"
    temp = await ctx.send(file=discord.File(path))
    await temp.delete(delay = 10)


@bot.command()
async def checkMission(ctx):
    CH = ctx.channel
    for channel in ctx.guild.channels:
        if channel.id == 955681106512973824 or channel.id == 944042644726759424:
            CH = channel
    wb = Workbook()
    member_list_index = ['display_name', 'name', 'discriminator', 'joined_at', 'role']
    ws = wb.active
    ws.append(member_list_index)
    await ctx.message.delete()
    li = []
    nowDate = dt.datetime.now()
    threshold = nowDate - dt.timedelta(days=15)
    async for message in CH.history(after = threshold, oldest_first = False):
        li.append(message.author.display_name)
    if ctx.author.display_name in li:
        temp = await ctx.send(ctx.author.display_name + "님은 미션을 완료하셨습니다. 안심해주세요!")
        await temp.delete(delay = 10)
    else:
        temp = await ctx.send(ctx.author.display_name + "님은 아직 미션을 진행하지 않으셨습니다. 미션을 진행해주세요!")
        await temp.delete(delay = 10)

@bot.command(pass_context=True, aliases = ["leave_sub", "join_sub"])
async def passcommand(ctx):
    pass

bot.run('OTQzMzU3MjcyODQxNjcwNjk3.Ygx37A.KYa6pk38l3xzN0mc4tT3IvXdCCE') #토큰
